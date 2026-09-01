#!/usr/bin/env python3

"""
Gemaakt door Thijs Vorstenburg en Ronald Koenis.

Het script is gemaakt om een MDTO-SIP te maken van 2 raadsvergaderingsbestanden die op een bronmap staan.
zodat het is in te lezen in bv MAIS-Flexis.

Input bestaat uit
- de YML-bestanden van de vergadering, dat vormt de basis
- de scans vanuit 'het analoge archief'.
- een excel met informatie over gescande documenten die via een id zijn gelinkt aan de YML.

Gebruik:
    zorg voor een lege map voor de uitvoer.
    python3 MaakMDTOuitYML.py --bronmap input --uitvoermap output --documentenlijst input/Documentenlijst.xlsx

De parameters kunnen ook via de omgevingsvariabele YML_BRONMAP, YML_UITVOERMAP en YML_DOCUMENTENLIJST worden ingesteld.

Prompt:
Maak een python script dat alle .yml-bestanden leest (gebruik de voorbeeldbestanden uit de bronnen) van een in te stellen bronlocatie, inclusief all submappen. 
Voor elk yml-bestand:
Toon van de meeting de eigenschappen meeting id, naam (staat in attributes met id 1), locatie (staat in attribute met id 50) en de datum en starttijd (staat in start_date).
En voor elk agenda-item:
Het id, het agendapunt (staat in de title_prefix), de naam (staat in attributes met id 1), de startseconde (staat in start_offset) en eindseconde (staat in end_offset)
Doe dit niet voor agenda-items
En voor elk document van het agenda-item:
Het id, de titel, de wijzigingsdatum, de categorie (staat in de eerste types met een ingevulde value) en de bestandsnaam.

"""

# Afhankelijkheden:

from __future__ import annotations
import argparse
import os
import re
import shutil
import sys
import tempfile
import mdto
from mdto.gegevensgroepen import *
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, TextIO, Optional
from urllib.parse import unquote, urlparse
from urllib.request import Request, urlopen

import yaml
from openpyxl import load_workbook

#De mdto-elementen die voor alles gelden:
id_uitgever = "GSV-RANU-pilot-edepot"

archiefvormer = VerwijzingGegevens(
    verwijzingNaam="gemeente Stichtse Vecht",
    verwijzingIdentificatie=IdentificatieGegevens("gm1904","TOOI register gemeenten compleet")
    ) 

#waardering is altijd blijvend te bewaren
waardering = BegripGegevens(
    begripLabel="Blijvend te bewaren",
    begripCode="B",
    begripBegrippenlijst=VerwijzingGegevens("Begrippenlijst Waarderingen MDTO")
    ) 

informatiecategorie = BegripGegevens(
    begripLabel="Agenda, verslag en besluitenlijst van bestuurlijke besluitvorming - Verwerkt",
    begripCode="19.1.6",
    begripBegrippenlijst=VerwijzingGegevens("Selectielijst gemeenten en intergemeentelijke organen 2017")
    )

informatiecategorieMotie = BegripGegevens(
    begripLabel="Adhesiebetuiging en/of motie - Ingewilligd",
    begripCode="6.1.4",
    begripBegrippenlijst=VerwijzingGegevens("Selectielijst gemeenten en intergemeentelijke organen 2017")
    )

# vooralsnog Geen beperkingen, TODO: aanpassen formatting en toevoegen andere attributen 
beperkingType = BegripGegevens("Geen beperking", VerwijzingGegevens("Begrippenlijst BeperkingGebruikTypeLijst MDTO"))
beperkingGebruik = BeperkingGebruikGegevens(beperkingGebruikType=beperkingType)



@dataclass(frozen=True)
class Scanverwijzing:
    """Verwijzing vanuit de documentenlijst naar een te kopiëren scan."""

    bronbestand: Path
    originele_bestandsnaam: str


@dataclass
class Documentstatistieken:
    """Tellingen en deduplicatie voor documenten van één meeting."""

    gekopieerd: int = 0
    gedownload: int = 0
    verwerkte_ids: set[str] = field(default_factory=set)


@dataclass
class Volgnummergenerator:
    """Reik unieke MDTO-namen uit binnen één volledige uitvoer."""

    volgend_nummer: int = 1

    def kandidaat(self) -> str:
        """Geef het volgende nummer zonder het al definitief uit te reiken."""
        return f"NL-BKLVV_1820_{self.volgend_nummer:04d}"

    def nieuw(self) -> str:
        naam = self.kandidaat()
        self.volgend_nummer += 1
        return naam


DOWNLOAD_TIMEOUT_SECONDEN = 60


def verwijder_ongeldige_tekens(tekst: str) -> tuple[str, int]:
    """Verwijder control-tekens die volgens YAML niet zijn toegestaan."""
    geldige_tekst = "".join(
        teken
        for teken in tekst
        if teken in "\t\n\r"
        or 0x20 <= ord(teken) <= 0x7E
        or ord(teken) >= 0xA0
    )
    return geldige_tekst, len(tekst) - len(geldige_tekst)


def lees_yml_bestanden(bronmap: Path) -> list[tuple[Path, Any]]:
    """Lees alle .yml-bestanden onder *bronmap*, alfabetisch gesorteerd."""
    if not bronmap.exists():
        raise FileNotFoundError(f"Bronmap bestaat niet: {bronmap}")
    if not bronmap.is_dir():
        raise NotADirectoryError(f"Bronlocatie is geen map: {bronmap}")

    resultaten: list[tuple[Path, Any]] = []
    for bestand in sorted(bronmap.rglob("*.yml")):
        try:
            # utf-8-sig verwerkt zowel gewone UTF-8 als de BOM in de voorbeelden.
            with bestand.open("r", encoding="utf-8-sig") as invoer:
                ruwe_tekst = invoer.read()
            yaml_tekst, aantal_verwijderd = verwijder_ongeldige_tekens(ruwe_tekst)
            if aantal_verwijderd:
                print(
                    f"Waarschuwing: {bestand}: {aantal_verwijderd} "
                    "ongeldig(e) control-teken(s) verwijderd",
                    file=sys.stderr,
                )
            inhoud = yaml.safe_load(yaml_tekst)
        except (OSError, UnicodeError, yaml.YAMLError) as fout:
            print(
                f"Waarschuwing: {bestand} kon niet worden gelezen: {fout}",
                file=sys.stderr,
            )
            continue
        resultaten.append((bestand, inhoud))
    return resultaten


def normaliseer_document_id(waarde: Any) -> str | None:
    """Normaliseer een Excel- of YAML-id naar een vergelijkbare tekenreeks."""
    if waarde is None or isinstance(waarde, bool):
        return None
    if isinstance(waarde, int):
        return str(waarde)
    if isinstance(waarde, float):
        return str(int(waarde)) if waarde.is_integer() else str(waarde)
    tekst = str(waarde).strip()
    if not tekst:
        return None
    if tekst.endswith(".0") and tekst[:-2].isdigit():
        return tekst[:-2]
    return tekst


def normaliseer_kolomnaam(waarde: Any) -> str:
    """Normaliseer een Excel-kolomnaam voor hoofdletterongevoelige matching."""
    return str(waarde or "").strip().lower().replace(" ", "_")


def lees_documentenlijst(excelpad: Path) -> dict[str, Scanverwijzing]:
    """Lees document-id, scanlocatie en scannaam uit een Excel-werkboek."""
    if not excelpad.exists():
        raise FileNotFoundError(f"Documentenlijst bestaat niet: {excelpad}")
    if not excelpad.is_file():
        raise OSError(f"Documentenlijst is geen bestand: {excelpad}")

    werkboek = load_workbook(excelpad, read_only=True, data_only=True)
    verwijzingen: dict[str, Scanverwijzing] = {}
    blad_met_kolommen = False
    try:
        for werkblad in werkboek.worksheets:
            rijen = iter(werkblad.iter_rows(values_only=True))
            kolommen: dict[str, int] | None = None

            # Zoek de kopregel in de eerste 50 rijen; lege titelregels zijn toegestaan.
            for _, rij in zip(range(50), rijen):
                kandidaat = {
                    normaliseer_kolomnaam(waarde): index
                    for index, waarde in enumerate(rij)
                    if heeft_waarde(waarde)
                }
                locatiekolom = kandidaat.get("scanlocatie", kandidaat.get("spanlocatie"))
                if "document_id" in kandidaat and locatiekolom is not None and "scannaam" in kandidaat:
                    kolommen = {
                        "document_id": kandidaat["document_id"],
                        "scanlocatie": locatiekolom,
                        "scannaam": kandidaat["scannaam"],
                    }
                    blad_met_kolommen = True
                    break

            if kolommen is None:
                continue

            for rij in rijen:
                def cel(kolom: str) -> Any:
                    index = kolommen[kolom]
                    return rij[index] if index < len(rij) else None

                document_id = normaliseer_document_id(cel("document_id"))
                if document_id is None:
                    continue
                scanlocatie = cel("scanlocatie")
                scannaam = cel("scannaam")
                if not heeft_waarde(scanlocatie) or not heeft_waarde(scannaam):
                    continue

                locatie = Path(str(scanlocatie).strip()).expanduser()
                if not locatie.is_absolute():
                    locatie = excelpad.parent / locatie
                originele_naam = Path(str(scannaam).strip()).name
                verwijzingen.setdefault(
                    document_id,
                    Scanverwijzing(
                        bronbestand=locatie / originele_naam,
                        originele_bestandsnaam=originele_naam,
                    ),
                )
    finally:
        werkboek.close()

    if not blad_met_kolommen:
        raise ValueError(
            "Geen werkblad gevonden met document_id, scanlocatie/spanlocatie "
            "en scannaam"
        )
    return verwijzingen


def waarde_met_id(items: Any, gezocht_id: int, sleutel: str = "value") -> Any:
    """Haal *sleutel* op uit het eerste item met het opgegeven id."""
    if not isinstance(items, list):
        return None
    for item in items:
        if isinstance(item, dict) and item.get("id") == gezocht_id:
            return item.get(sleutel)
    return None


def heeft_waarde(waarde: Any) -> bool:
    """Bepaal of een veld een niet-lege waarde bevat."""
    return waarde is not None and str(waarde).strip() != ""


def eerste_gevulde_waarde(items: Any) -> Any:
    """Geef de eerste niet-lege value uit een lijst met mappings terug."""
    if not isinstance(items, list):
        return None
    for item in items:
        if isinstance(item, dict) and heeft_waarde(item.get("value")):
            return item["value"]
    return None


def agenda_attributen(agenda_item: dict[str, Any]) -> Any:
    """Geef de attributen uit type_data van een agenda-item terug."""
    type_data = agenda_item.get("type_data")
    return type_data.get("attributes") if isinstance(type_data, dict) else None


def agenda_item_kop(agenda_item: dict[str, Any]) -> tuple[Any, Any]:
    """Geef het agendapunt en de naam van een agenda-item terug."""
    type_data = agenda_item.get("type_data")
    agendapunt = type_data.get("title_prefix") if isinstance(type_data, dict) else None
    naam = waarde_met_id(agenda_attributen(agenda_item), 1)
    return agendapunt, naam


def is_bestandsdocument(document: dict[str, Any]) -> bool:
    """Controleer of minstens één documentversie type 'file' heeft."""
    versions = document.get("versions")
    if not isinstance(versions, list):
        return False
    return any(
        isinstance(version, dict)
        and str(version.get("type", "")).strip().lower() == "file"
        for version in versions
    )


def eerste_bestandsnaam(document: dict[str, Any]) -> Any:
    """Geef de eerste bestandsnaam van een versie met type 'file' terug."""
    versions = document.get("versions")
    if not isinstance(versions, list):
        return None
    for version in versions:
        if (
            isinstance(version, dict)
            and str(version.get("type", "")).strip().lower() == "file"
            and heeft_waarde(version.get("file_name"))
        ):
            return version["file_name"]
    return None


def iter_agenda_items(items: Any, niveau: int = 0):
    """Doorloop alleen benoemde agenda-items, inclusief benoemde subitems."""
    if not isinstance(items, list):
        return
    for item in items:
        if not isinstance(item, dict):
            continue
        agendapunt, naam = agenda_item_kop(item)
        tonen = heeft_waarde(agendapunt) or heeft_waarde(naam)
        if tonen:
            yield item, niveau
        # Geldige kinderen van een overgeslagen leeg item blijven zichtbaar.
        kindniveau = niveau + 1 if tonen else niveau
        yield from iter_agenda_items(item.get("agenda_items"), kindniveau)


def maak_kopienaam(document_id: str, originele_bestandsnaam: str) -> str:
    """Maak '<id>_<naam>' en pas de voorgeschreven tekenopschoning toe."""
    bestandsnaam = Path(originele_bestandsnaam).name
    laatste_punt = bestandsnaam.rfind(".")
    heeft_extensie = 0 < laatste_punt < len(bestandsnaam) - 1

    if heeft_extensie:
        stam = bestandsnaam[:laatste_punt].replace(".", "")
        extensie = bestandsnaam[laatste_punt + 1 :]
    else:
        stam = bestandsnaam.replace(".", "")
        extensie = ""

    vertaaltabel = str.maketrans("", "", ",+()-")
    stam = stam.translate(vertaaltabel).replace(" ", "_")
    extensie = extensie.translate(vertaaltabel).replace(" ", "_")
    opgeschoond = f"{stam}.{extensie}" if extensie else stam
    return re.sub(r"_+", "_", f"{document_id}_{opgeschoond}")


def bestandsnaam_voor_document(
    document: dict[str, Any], verwijzing: Scanverwijzing | None
) -> str:
    """Bepaal de originele bestandsnaam uit Excel, YAML of de document-URL."""
    if verwijzing is not None and heeft_waarde(verwijzing.originele_bestandsnaam):
        return verwijzing.originele_bestandsnaam
    yaml_naam = eerste_bestandsnaam(document)
    if heeft_waarde(yaml_naam):
        return Path(str(yaml_naam)).name
    url = document.get("url")
    if heeft_waarde(url):
        url_naam = Path(unquote(urlparse(str(url)).path)).name
        if heeft_waarde(url_naam):
            return url_naam
    return "document"


def download_document(url: str, doelbestand: Path) -> None:
    """Download een HTTP(S)-document veilig via een tijdelijk bestand."""
    parsed = urlparse(url)
    if parsed.scheme.lower() not in {"http", "https"}:
        raise ValueError(f"niet-ondersteund URL-schema: {parsed.scheme or 'ontbreekt'}")

    verzoek = Request(url, headers={"User-Agent": "Maak-MDTO-YML-import/1.0"})
    tijdelijk_pad: Path | None = None
    try:
        with urlopen(verzoek, timeout=DOWNLOAD_TIMEOUT_SECONDEN) as antwoord:
            definitieve_url = antwoord.geturl()
            if urlparse(definitieve_url).scheme.lower() not in {"http", "https"}:
                raise ValueError("download is doorgestuurd naar een niet-HTTP(S)-URL")
            with tempfile.NamedTemporaryFile(
                mode="wb",
                dir=doelbestand.parent,
                prefix=".download-",
                suffix=".tmp",
                delete=False,
            ) as tijdelijk_bestand:
                tijdelijk_pad = Path(tijdelijk_bestand.name)
                shutil.copyfileobj(antwoord, tijdelijk_bestand)
        tijdelijk_pad.replace(doelbestand)
        tijdelijk_pad = None
    finally:
        if tijdelijk_pad is not None:
            tijdelijk_pad.unlink(missing_ok=True)


def schrijf_document_mdto_xml(
    verwijzing: VerwijzingGegevens(),
    document: dict[str, Any], 
    doelbestand: Path, 
    document_id: str,
    mdto_naam: str,
    dekking_in_tijd: DekkingInTijdGegevens,
) -> None:
    """Schrijf de sidecar na overdracht en meld een eventuele schrijffout."""

    titel = document.get("title")
    documentnaam = "" if titel is None else str(titel)
    
    id = id_uitgever + ".D." + document_id
    bestandInformatieobjectIdentificatieGegegevens = IdentificatieGegevens(id, "gemeente Stichtse Vecht")

    # maak informatieobject voor het bestand:
    informatieobject = Informatieobject (
        identificatie = bestandInformatieobjectIdentificatieGegegevens, 
        naam = documentnaam,
        waardering = waardering,
        informatiecategorie = informatiecategorie,
        #todo
        # informatiecategorie = informatiecategorieMotie if str(rij["doc.classificatie"]).strip().lower() == "motie"  else informatiecategorie,
        archiefvormer = archiefvormer,
        taal = "nl",
        dekkingInTijd = dekking_in_tijd,
        beperkingGebruik = beperkingGebruik, 
        classificatie = BegripGegevens("Vergadermedia" if "media" in document_id else "Document" , VerwijzingGegevens("Begrippenlijst Archiefeenheidsoorten MAIS-Flexis")),
        aggregatieniveau = BegripGegevens("Archiefstuk", VerwijzingGegevens("Begrippenlijst Aggregatieniveaus MDTO")),
        #todo  classificatie = BegripGegevens(str(rij["doc.classificatie"]).strip(), VerwijzingGegevens("Begrippenlijst TODO")),
        isOnderdeelVan = verwijzing
        ) 

    metadata_bestand = doelbestand.parent / f"{mdto_naam}.mdto.xml"
    try:
        informatieobject.save(metadata_bestand)
    except OSError as fout:
        print(
            f"Waarschuwing: MDTO-bestand voor document {document_id} kon niet "
            f"worden geschreven: {fout}",
            file=sys.stderr,
        )

    # en daarna de xml voor het bestand zelf: 
    representatie = VerwijzingGegevens( 
        verwijzingNaam=informatieobject.naam, 
        verwijzingIdentificatie=bestandInformatieobjectIdentificatieGegegevens 
        ) 

    # Mimetype (True) geeft error met aanmaken van bepaalde files, PRONOM (false) heeft daar geen last van. Sowieso is Pronom een betere identificatie dan mimetype 
    bestandobject = Bestand.from_file(
        file_or_filename=doelbestand,
        isRepresentatieVan=representatie, 
        use_mimetype=False, 
        ) 
    
    metadata_bestand = doelbestand.parent / f"{mdto_naam}.bestand.mdto.xml"
    try:
        bestandobject.save(metadata_bestand) 
    except OSError as fout:
        print(
            f"Waarschuwing: MDTO-bestand.mdto.xml voor document {document_id} kon niet "
            f"worden geschreven: {fout}",
            file=sys.stderr,
        )


def verwerk_documentbestand(
    verwijzingGeg: VerwijzingGegevens(),
    document: dict[str, Any],
    documentenlijst: dict[str, Scanverwijzing],
    doelmap: Path,
    statistieken: Documentstatistieken,
    volgnummers: Volgnummergenerator,
    dekking_in_tijd: DekkingInTijdGegevens,
) -> None:
    """Kopieer of download één document dat door schrijf_document is verwerkt."""
    document_id = normaliseer_document_id(document.get("id"))
    if document_id is None or document_id in statistieken.verwerkte_ids:
        return
    statistieken.verwerkte_ids.add(document_id)
    verwijzing = documentenlijst.get(document_id)

    originele_naam = bestandsnaam_voor_document(document, verwijzing)
    # Reik het nummer pas definitief uit nadat kopiëren/downloaden is gelukt.
    # Zo veroorzaken niet-verwerkte documenten geen lege map of gat in de reeks.
    mdto_naam = volgnummers.kandidaat()
    documentmap = doelmap / mdto_naam
    documentmap.mkdir(parents=True, exist_ok=True)
    extensie = Path(originele_naam).suffix
    doelbestand = documentmap / f"{mdto_naam}{extensie}"

    if verwijzing is not None and verwijzing.bronbestand.is_file():
        try:
            shutil.copy2(verwijzing.bronbestand, doelbestand)
        except OSError as fout:
            print(
                f"Waarschuwing: document {document_id} kon niet vanuit "
                f"de scanlocatie worden gekopieerd: {fout}; download wordt geprobeerd",
                file=sys.stderr,
            )
        else:
            volgnummers.nieuw()
            statistieken.gekopieerd += 1
            schrijf_document_mdto_xml(
                verwijzingGeg,
                document,
                doelbestand,
                document_id,
                mdto_naam,
                dekking_in_tijd,
            )
            return

    url = document.get("url")
    if not heeft_waarde(url):
        reden = "staat niet in Excel"
        if verwijzing is not None:
            reden = f"ontbreekt op scanlocatie {verwijzing.bronbestand}"
        print(
            f"Waarschuwing: document {document_id} {reden} en heeft geen URL",
            file=sys.stderr,
        )
        doelbestand.unlink(missing_ok=True)
        documentmap.rmdir()
        return

    try:
        download_document(str(url).strip(), doelbestand)
    except (OSError, ValueError) as fout:
        print(
            f"Waarschuwing: document {document_id} kon niet worden "
            f"gedownload via {url}: {fout}",
            file=sys.stderr,
        )
        doelbestand.unlink(missing_ok=True)
        documentmap.rmdir()
        return
    volgnummers.nieuw()
    statistieken.gedownload += 1
    schrijf_document_mdto_xml(
        verwijzingGeg,
        document,
        doelbestand,
        document_id,
        mdto_naam,
        dekking_in_tijd,
    )


def schrijf_document(
    verwijzing: VerwijzingGegevens(),
    document: dict[str, Any],
    inspringing: str,
    uitvoer: TextIO,
    documentenlijst: dict[str, Scanverwijzing],
    doelmap: Path,
    statistieken: Documentstatistieken,
    volgnummers: Volgnummergenerator,
    dekking_in_tijd: DekkingInTijdGegevens,
) -> None:
    """Schrijf één document en kopieer of download het bijbehorende bestand."""
    categorie = eerste_gevulde_waarde(document.get("types")) or "-"
    bestandsnaam = eerste_bestandsnaam(document) or "-"
    print(f"{inspringing}Document", file=uitvoer)
    print(f"{inspringing}  id:              {document.get('id', '-')}", file=uitvoer)
    print(f"{inspringing}  titel:           {document.get('title', '-')}", file=uitvoer)
    print(
        f"{inspringing}  wijzigingsdatum: {document.get('last_modified', '-')}",
        file=uitvoer,
    )
    print(f"{inspringing}  categorie:       {categorie}", file=uitvoer)
    print(f"{inspringing}  bestandsnaam:    {bestandsnaam}", file=uitvoer)
    verwerk_documentbestand(
        verwijzing,
        document,
        documentenlijst,
        doelmap,
        statistieken,
        volgnummers,
        dekking_in_tijd,
    )


def schrijf_media(
    vergadering_verwijzing: VerwijzingGegevens,
    meeting_id: str,
    mediatype: str,
    volgorde: int,
    media_item: dict[str, Any],
    uitvoer: TextIO,
    documentenlijst: dict[str, Scanverwijzing],
    doelmap: Path,
    statistieken: Documentstatistieken,
    volgnummers: Volgnummergenerator,
    dekking_in_tijd: DekkingInTijdGegevens,
) -> None:
    """Verwerk meetingmedia met dezelfde bestandslogica als een document."""
    bestandsnaam = media_item.get("filename")
    if not heeft_waarde(bestandsnaam):
        print(
            f"Waarschuwing: {mediatype} {volgorde} van meeting {meeting_id} "
            "heeft geen bestandsnaam en wordt overgeslagen",
            file=sys.stderr,
        )
        return

    download_url = media_item.get("download")
    if heeft_waarde(download_url) and str(download_url).startswith("//"):
        download_url = "https:" + str(download_url)

    media_id = f"{meeting_id}-media-{mediatype}-{volgorde}"
    media_document = {
        "id": media_id,
        "title": str(bestandsnaam),
        "url": download_url,
        "types": [{"value": mediatype}],
        "versions": [{"type": "file", "file_name": str(bestandsnaam)}],
    }

    print("  Media", file=uitvoer)
    print(f"    id:              {media_id}", file=uitvoer)
    print(f"    type:            {mediatype}", file=uitvoer)
    print(f"    bestandsnaam:    {bestandsnaam}", file=uitvoer)
    print(f"    download:        {download_url or '-'}", file=uitvoer)

    verwerk_documentbestand(
        vergadering_verwijzing,
        media_document,
        documentenlijst,
        doelmap,
        statistieken,
        volgnummers,
        dekking_in_tijd,
    )


def schrijf_meetingmedia(
    meeting: dict[str, Any],
    vergadering_verwijzing: VerwijzingGegevens,
    uitvoer: TextIO,
    documentenlijst: dict[str, Scanverwijzing],
    doelmap: Path,
    statistieken: Documentstatistieken,
    volgnummers: Volgnummergenerator,
    dekking_in_tijd: DekkingInTijdGegevens,
) -> None:
    """Verwerk alle audio en video die rechtstreeks bij de meeting horen."""
    meeting_id = normaliseer_document_id(meeting.get("id"))
    media = meeting.get("media")
    if meeting_id is None or not isinstance(media, dict):
        return

    for mediatype in ("audio", "video"):
        media_items = media.get(mediatype)
        if not isinstance(media_items, list):
            continue
        for volgorde, media_item in enumerate(media_items, start=1):
            if isinstance(media_item, dict):
                schrijf_media(
                    vergadering_verwijzing,
                    meeting_id,
                    mediatype,
                    volgorde,
                    media_item,
                    uitvoer,
                    documentenlijst,
                    doelmap,
                    statistieken,
                    volgnummers,
                    dekking_in_tijd,
                )


def schrijf_agendapunt_mdto_xml(
    isonderdeelvan : VerwijzingGegevens,
    agenda_item: dict[str, Any], 
    agendapuntnaam, 
    doelmap: Path,
    volgnummers: Volgnummergenerator,
    dekking_in_tijd: DekkingInTijdGegevens,
) -> VerwijzingGegevens | None:
    """Schrijf de naam van een agenda-item naar een eigen MDTO-sidecar."""
    agenda_id = normaliseer_document_id(agenda_item.get("id"))
    if agenda_id is None:
        print(
            "Waarschuwing: voor een agenda-item zonder id kan geen MDTO-bestand "
            "worden geschreven",
            file=sys.stderr,
        )
        return None

    title_prefix, naam = agenda_item_kop(agenda_item)
    title_prefix = "" if title_prefix is None else str(title_prefix)
    title_prefix = title_prefix.replace(".", "").strip()
    # Voorkom dat een afwijkende title_prefix onbedoeld een submap vormt.
    title_prefix = title_prefix.replace("/", "_").replace("\\", "_")

    #mdto opbouwen:
    agendapuntId = id_uitgever + ".A." + agenda_id
    agendapuntnaam = f"Agendapunt {title_prefix}: " + agendapuntnaam
    print(f"Nieuw agendapunt {agendapuntnaam}")

    # maak identificatiekenmerk element
    agendapuntInformatieobjectIdentificatieGegegevens = IdentificatieGegevens(agendapuntId, "gemeente Stichtse Vecht")

    # maak agendaobject op basis van deze gegevens
    informatieobject = Informatieobject (
        identificatie = agendapuntInformatieobjectIdentificatieGegegevens,
        naam = agendapuntnaam,
        waardering = waardering,
        informatiecategorie = informatiecategorie,
        archiefvormer = archiefvormer,
        taal = "nl",
        dekkingInTijd = dekking_in_tijd,
        beperkingGebruik = beperkingGebruik,
        classificatie = BegripGegevens("Agendapunt", VerwijzingGegevens("Begrippenlijst Archiefeenheidsoorten MAIS-Flexis")),
        aggregatieniveau = BegripGegevens("Dossier", VerwijzingGegevens("Begrippenlijst Aggregatieniveaus MDTO")),
        isOnderdeelVan = isonderdeelvan
        )

    metadata_bestand = doelmap / f"{volgnummers.nieuw()}.mdto.xml"
    inhoud = "" if naam is None else str(naam)
    try:
        informatieobject.save(metadata_bestand)
    except OSError as fout:
        print(
            f"Waarschuwing: MDTO-bestand voor agenda-item {agenda_id} kon niet "
            f"worden geschreven: {fout}",
            file=sys.stderr,
        )
        return None

    AgendapuntVerwijzingGegevens = VerwijzingGegevens( 
        verwijzingNaam = agendapuntnaam,
        verwijzingIdentificatie = agendapuntInformatieobjectIdentificatieGegegevens 
    )
    
    return AgendapuntVerwijzingGegevens


def schrijf_vergadering_mdto_xml(
    meeting: dict[str, Any],
    start_date: Any,
    doelmap: Path,
    volgnummers: Volgnummergenerator,
    dekking_in_tijd: DekkingInTijdGegevens,
) -> VerwijzingGegevens | None:
    """Schrijf de meetingnaam naar '<id>_<naam>_<jjjjmmdd>.mdto_xml'."""
    meeting_id = normaliseer_document_id(meeting.get("id"))
    meetingnaam = waarde_met_id(meeting.get("attributes"), 1)
    meetingnaam = str(meetingnaam).strip() if heeft_waarde(meetingnaam) else ""

    if meeting_id is None or not meetingnaam or not heeft_waarde(start_date):
        print(
            "Waarschuwing: voor een meeting zonder id, naam of start_date kan "
            "geen MDTO-bestand worden geschreven",
            file=sys.stderr,
        )
        return None

    datumtekst = str(start_date).strip().split(maxsplit=1)[0]
    try:
        datum_voor_naam = date.fromisoformat(datumtekst).strftime("%Y%m%d")
    except ValueError:
        print(
            f"Waarschuwing: MDTO-bestand voor meeting {meeting_id} kon niet "
            f"worden geschreven; ongeldige start_date: {start_date}",
            file=sys.stderr,
        )
        return None

    print(f"Nieuwe vergadering {meetingnaam} op " + date.fromisoformat(datumtekst).strftime("%Y-%m-%d"))
    # maak identificatiekenmerk element
    vergaderingId = id_uitgever+".V."+meeting_id
    vergaderingInformatieobjectIdentificatieGegegevens = IdentificatieGegevens(vergaderingId, "gemeente Stichtse Vecht")
    
    # vergaderdatum
    # TODO: de aanvangstijd ergens aan toevoegen.
    datum_string = date.fromisoformat(datumtekst).strftime("%Y-%m-%d")

    # maak vergaderobject op basis van deze gegevens
    informatieobject = Informatieobject (
        identificatie = vergaderingInformatieobjectIdentificatieGegegevens,
        naam = meetingnaam + " " + datum_string,
        waardering = waardering,
        informatiecategorie = informatiecategorie,
        archiefvormer = archiefvormer,
        taal = "nl",
        dekkingInTijd = dekking_in_tijd,
        beperkingGebruik = beperkingGebruik,
        classificatie = BegripGegevens("Vergaderagenda", VerwijzingGegevens("Begrippenlijst Archiefeenheidsoorten MAIS-Flexis")),
        aggregatieniveau = BegripGegevens("Dossier", VerwijzingGegevens("Begrippenlijst Aggregatieniveaus MDTO"))
        )

    metadata_bestand = doelmap / f"{volgnummers.nieuw()}.mdto.xml"
    try:
        informatieobject.save(metadata_bestand)
    except OSError as fout:
        print(
            f"Waarschuwing: MDTO-bestand voor meeting {meeting_id} kon niet "
            f"worden geschreven: {fout}",
            file=sys.stderr,
        )
        return None

    VergaderingVerwijzingGegevens = VerwijzingGegevens( 
            verwijzingNaam = meetingnaam + " " + datum_string,
            verwijzingIdentificatie = vergaderingInformatieobjectIdentificatieGegegevens 
        )

    return VergaderingVerwijzingGegevens


def schrijf_agenda_item(
    isonderdeelvan : VerwijzingGegevens,
    agenda_item: dict[str, Any],
    niveau: int,
    uitvoer: TextIO,
    documentenlijst: dict[str, Scanverwijzing],
    doelmap: Path,
    statistieken: Documentstatistieken,
    volgnummers: Volgnummergenerator,
    dekking_in_tijd: DekkingInTijdGegevens,
) -> None:
    """Schrijf een agenda-item en alle direct gekoppelde documenten."""
    inspringing = "  " * (niveau + 1)
    agendapunt, naam = agenda_item_kop(agenda_item)
    agendapunt = agendapunt if heeft_waarde(agendapunt) else "-"
    naam = naam if heeft_waarde(naam) else "-"
    startseconde = agenda_item.get("start_offset")
    eindseconde = agenda_item.get("end_offset")
    startseconde = "-" if startseconde in (None, "") else startseconde
    eindseconde = "-" if eindseconde in (None, "") else eindseconde

    print(f"{inspringing}Agenda-item", file=uitvoer)
    print(f"{inspringing}  id:            {agenda_item.get('id', '-')}", file=uitvoer)
    print(f"{inspringing}  agendapunt:    {agendapunt}", file=uitvoer)
    print(f"{inspringing}  naam:          {naam}", file=uitvoer)
    print(f"{inspringing}  startseconde:  {startseconde}", file=uitvoer)
    print(f"{inspringing}  eindseconde:   {eindseconde}", file=uitvoer)

    AgendapuntVerwijzingGegevens = schrijf_agendapunt_mdto_xml(
        isonderdeelvan,
        agenda_item,
        naam,
        doelmap,
        volgnummers,
        dekking_in_tijd,
    )

    documenten = agenda_item.get("documents")
    if isinstance(documenten, list):
        for document in documenten:
            if isinstance(document, dict) and is_bestandsdocument(document):
                schrijf_document(
                    AgendapuntVerwijzingGegevens,
                    document,
                    inspringing + "  ",
                    uitvoer,
                    documentenlijst,
                    doelmap,
                    statistieken,
                    volgnummers,
                    dekking_in_tijd,
                )


def schrijf_meeting(
    bestand: Path,
    inhoud: Any,
    uitvoer: TextIO,
    documentenlijst: dict[str, Scanverwijzing],
    doelmap: Path,
    statistieken: Documentstatistieken,
    volgnummers: Volgnummergenerator,
) -> None:
    """Schrijf meeting, agenda-items en documenten uit een Notubiz-export."""
    if not isinstance(inhoud, dict):
        print(f"Bestand: {bestand}", file=uitvoer)
        print("YAML-hoofdstructuur is geen mapping", file=uitvoer)
        return

    naam = waarde_met_id(inhoud.get("attributes"), 1) or "-"
    locatie = waarde_met_id(inhoud.get("attributes"), 50) or "-"
    planning = inhoud.get("plannings") or []
    start_date = planning[0].get("start_date") if planning else None
    datum, starttijd = "-", "-"
    if start_date:
        delen = str(start_date).split(maxsplit=1)
        datum = delen[0]
        starttijd = delen[1] if len(delen) == 2 else "-"

    print(f"Bestand: {bestand}", file=uitvoer)
    print("Meeting", file=uitvoer)
    print(f"  meeting id: {inhoud.get('id', '-')}", file=uitvoer)
    print(f"  naam:       {naam}", file=uitvoer)
    print(f"  locatie:    {locatie}", file=uitvoer)
    print(f"  datum:      {datum}", file=uitvoer)
    print(f"  starttijd:  {starttijd}", file=uitvoer)

    if not heeft_waarde(start_date):
        print(
            "Waarschuwing: zonder start_date kan geen DekkingInTijd-element "
            "worden gemaakt",
            file=sys.stderr,
        )
        return
    datumtekst = str(start_date).strip().split(maxsplit=1)[0]
    try:
        datum_string = date.fromisoformat(datumtekst).strftime("%Y-%m-%d")
    except ValueError:
        print(
            f"Waarschuwing: ongeldige start_date: {start_date}",
            file=sys.stderr,
        )
        return

    vergadering_dekking_in_tijd = DekkingInTijdGegevens(
        dekkingInTijdType=BegripGegevens(
            "vergaderdatum", VerwijzingGegevens("Begrippenlijst TODO")
        ),
        dekkingInTijdBegindatum=datum_string,
    )

    VergaderingVerwijzingGegevens = schrijf_vergadering_mdto_xml(
        inhoud,
        start_date,
        doelmap,
        volgnummers,
        vergadering_dekking_in_tijd,
    )

    schrijf_meetingmedia(
        inhoud,
        VergaderingVerwijzingGegevens,
        uitvoer,
        documentenlijst,
        doelmap,
        statistieken,
        volgnummers,
        vergadering_dekking_in_tijd,
    )

    for agenda_item, niveau in iter_agenda_items(inhoud.get("agenda_items")):
        schrijf_agenda_item(
            VergaderingVerwijzingGegevens,
            agenda_item,
            niveau,
            uitvoer,
            documentenlijst,
            doelmap,
            statistieken,
            volgnummers,
            vergadering_dekking_in_tijd,
        )


def uitvoerpad(bronmap: Path, uitvoermap: Path, bronbestand: Path) -> Path:
    """Bepaal het .txt-uitvoerpad met behoud van de relatieve submappen."""
    relatief_pad = bronbestand.relative_to(bronmap)
    return (uitvoermap / relatief_pad).with_suffix(".txt")


def verwerk_bronmap(
    bronmap: Path, uitvoermap: Path, documentenlijstpad: Path
) -> int:
    """Schrijf meetings; schrijf_document verwerkt de bijbehorende bestanden."""
    if uitvoermap.is_dir():
        try:
            uitvoermap_is_niet_leeg = next(uitvoermap.iterdir(), None) is not None
        except OSError as fout:
            print(
                f"Waarschuwing: de inhoud van uitvoermap {uitvoermap} kon niet "
                f"worden gecontroleerd: {fout}",
                file=sys.stderr,
            )
            sys.exit("voortijdig einde")
        else:
            # Checking if the list is empty or not, blijkbaar staat in een lege map op mijn pc toch nog 1 onzichtbaar bestandje, vandaar:
            if len(os.listdir(uitvoermap)) > 1:
                print(
                    f"Waarschuwing: uitvoermap {uitvoermap} is niet leeg; ",
                    file=sys.stderr,
                )
                sys.exit("voortijdig einde")

    try:
        documentenlijst = lees_documentenlijst(documentenlijstpad)
    except (OSError, ValueError) as fout:
        print(f"Fout bij lezen van de documentenlijst: {fout}", file=sys.stderr)
        return 4

    try:
        documenten = lees_yml_bestanden(bronmap)
    except (FileNotFoundError, NotADirectoryError) as fout:
        print(f"Fout: {fout}", file=sys.stderr)
        return 2

    if not documenten:
        print(f"Geen leesbare .yml-bestanden gevonden in {bronmap}", file=sys.stderr)
        return 1

    aantal_geschreven = 0
    aantal_gekopieerd = 0
    aantal_gedownload = 0
    volgnummers = Volgnummergenerator()
    for bestand, inhoud in documenten:
        doelbestand = uitvoerpad(bronmap, uitvoermap, bestand)
        statistieken = Documentstatistieken()
        try:
            doelbestand.parent.mkdir(parents=True, exist_ok=True)
            with doelbestand.open("w", encoding="utf-8", newline="\n") as uitvoer:
                schrijf_meeting(
                    bestand.relative_to(bronmap),
                    inhoud,
                    uitvoer,
                    documentenlijst,
                    doelbestand.parent,
                    statistieken,
                    volgnummers,
                )
        except OSError as fout:
            print(
                f"Waarschuwing: uitvoer kon niet worden geschreven naar "
                f"{doelbestand}: {fout}",
                file=sys.stderr,
            )
            continue
        aantal_geschreven += 1
        aantal_gekopieerd += statistieken.gekopieerd
        aantal_gedownload += statistieken.gedownload

    if aantal_geschreven == 0:
        return 3

    print(
        f"{aantal_geschreven} uitvoerbestand(en) geschreven en "
        f"{aantal_gekopieerd} documentbestand(en) gekopieerd en "
        f"{aantal_gedownload} documentbestand(en) gedownload naar "
        f"{uitvoermap.resolve()}"
    )
    return 0


def parse_args() -> argparse.Namespace:
    standaard_bronmap = os.environ.get("YML_BRONMAP", "project_sources")
    standaard_uitvoermap = os.environ.get("YML_UITVOERMAP", "uitvoer")
    standaard_documentenlijst = os.environ.get("YML_DOCUMENTENLIJST", "Documentenlijst.xlsx")
    parser = argparse.ArgumentParser(
        description=(
            "Lees Notubiz-YAML en een Excel-documentenlijst, schrijf per bron "
            "een tekstbestand, kopieer gekoppelde scans en download ontbrekende "
            "documenten via de YAML-URL."
        )
    )
    parser.add_argument(
        "--bronmap",
        type=Path,
        default=Path(standaard_bronmap),
        help=f"Hoofdmap met .yml-bestanden (standaard: {standaard_bronmap!r})",
    )
    parser.add_argument(
        "--uitvoermap",
        type=Path,
        default=Path(standaard_uitvoermap),
        help=f"Hoofdmap voor .txt-uitvoer (standaard: {standaard_uitvoermap!r})",
    )
    parser.add_argument(
        "--documentenlijst",
        type=Path,
        default=Path(standaard_documentenlijst),
        help=(
            "Excel met document_id, scanlocatie/spanlocatie en scannaam "
            f"(standaard: {standaard_documentenlijst!r})"
        ),
    )
    return parser.parse_args()

def main() -> int:
    args = parse_args()
    return verwerk_bronmap(args.bronmap, args.uitvoermap, args.documentenlijst)


if __name__ == "__main__":
    raise SystemExit(main())

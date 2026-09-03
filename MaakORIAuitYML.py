#!/usr/bin/env python3
"""Maak ORI-A XML en verzamel bestanden uit Notubiz-YAML.

Het script leest recursief alle ``.yml``-bestanden uit een bronmap. Per meeting
wordt één ORI-A 1.0.1 XML-bestand gemaakt. Documenten en meetingmedia worden,
indien gewenst, uit een Excel-lijst gekopieerd of anders via hun URL gedownload.

Voorbeeld:

    python3 MaakORIAuitYML.py \
        --bronmap input \
        --uitvoermap output \
        --documentenlijst input/Documentenlijst.xlsx \
        --xsd ORI-A-XSD/ORI-A.xsd

De locaties kunnen ook worden ingesteld met ``YML_BRONMAP``,
``YML_UITVOERMAP``, ``YML_DOCUMENTENLIJST`` en ``ORIA_XSD``.
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse
from urllib.request import Request, urlopen
from xml.etree import ElementTree as ET

import yaml
from openpyxl import load_workbook


ORIA_NS = "https://ori-a.nl"
XSI_NS = "http://www.w3.org/2001/XMLSchema-instance"
SCHEMA_LOCATIE = (
    "https://ori-a.nl "
    "https://github.com/Regionaal-Archief-Rivierenland/"
    "ORI-A-XSD/releases/download/v1.0.1/ORI-A.xsd"
)
VERGADERSTUKTYPEN = "https://ori-a.nl/begrippenlijsten#vergaderstuktypes"
MEDIABRONTYPEN = "https://ori-a.nl/begrippenlijsten#mediabrontypes"
VERGADERINGSTYPEN = "https://ori-a.nl/begrippenlijsten#vergaderingstypes"
GEMEENTENREGISTER = (
    "https://identifier.overheid.nl/tooi/set/rwc_gemeenten_compleet/4"
)
DOWNLOAD_TIMEOUT_SECONDEN = 60
STANDAARD_PREFIX = "NL-BKLVV_1820"

ET.register_namespace("", ORIA_NS)
ET.register_namespace("xsi", XSI_NS)


@dataclass(frozen=True)
class Scanverwijzing:
    """Verwijzing vanuit Excel naar een lokaal scanbestand."""

    bronbestand: Path
    originele_bestandsnaam: str


@dataclass
class Statistieken:
    """Tellingen en deduplicatie voor één volledige verwerking."""

    xml_bestanden: int = 0
    gekopieerd: int = 0
    gedownload: int = 0
    overgeslagen: int = 0
    verwerkte_document_ids: set[str] = field(default_factory=set)


@dataclass
class Meetingnummering:
    """Maak namen met meeting- en objectvolgnummer van elk vier cijfers."""

    meetingvolgnummer: int
    prefix: str = STANDAARD_PREFIX
    objectvolgnummer: int = 1

    def kandidaat(self) -> str:
        return (
            f"{self.prefix}_{self.meetingvolgnummer:04d}_"
            f"{self.objectvolgnummer:04d}"
        )

    def nieuw(self) -> str:
        naam = self.kandidaat()
        self.objectvolgnummer += 1
        return naam


class _HTMLNaarTekst(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.delen: list[str] = []

    def handle_data(self, data: str) -> None:
        if data.strip():
            self.delen.append(data.strip())


def html_naar_tekst(waarde: Any) -> str | None:
    """Maak leesbare platte tekst van een eventueel HTML-veld."""
    if not heeft_waarde(waarde):
        return None
    parser = _HTMLNaarTekst()
    parser.feed(str(waarde))
    tekst = " ".join(parser.delen)
    return re.sub(r"\s+", " ", tekst).strip() or None


def heeft_waarde(waarde: Any) -> bool:
    return waarde is not None and str(waarde).strip() != ""


def normaliseer_id(waarde: Any) -> str | None:
    """Normaliseer numerieke Excel-/YAML-id's zonder een eventuele .0."""
    if not heeft_waarde(waarde) or isinstance(waarde, bool):
        return None
    tekst = str(waarde).strip()
    if re.fullmatch(r"[+-]?\d+\.0+", tekst):
        tekst = tekst.split(".", 1)[0]
    return tekst


def waarde_met_id(items: Any, gezocht_id: int, sleutel: str = "value") -> Any:
    if not isinstance(items, list):
        return None
    for item in items:
        if isinstance(item, dict) and item.get("id") == gezocht_id:
            return item.get(sleutel)
    return None


def eerste_gevulde_waarde(items: Any) -> Any:
    if not isinstance(items, list):
        return None
    for item in items:
        if isinstance(item, dict) and heeft_waarde(item.get("value")):
            return item["value"]
    return None


def verwijder_ongeldige_tekens(tekst: str) -> tuple[str, int]:
    """Verwijder tekens die niet zijn toegestaan in XML 1.0 en YAML."""
    geldig = "".join(
        teken
        for teken in tekst
        if teken in "\t\n\r"
        or 0x20 <= ord(teken) <= 0x7E
        or 0xA0 <= ord(teken) <= 0xD7FF
        or 0xE000 <= ord(teken) <= 0xFFFD
    )
    return geldig, len(tekst) - len(geldig)


def lees_yml_bestanden(bronmap: Path) -> list[tuple[Path, Any]]:
    """Lees alle .yml-bestanden in de bronmap en haar submappen."""
    if not bronmap.exists():
        raise FileNotFoundError(f"Bronmap bestaat niet: {bronmap}")
    if not bronmap.is_dir():
        raise NotADirectoryError(f"Bronlocatie is geen map: {bronmap}")

    resultaten: list[tuple[Path, Any]] = []
    for bestand in sorted(bronmap.rglob("*.yml")):
        try:
            ruwe_tekst = bestand.read_text(encoding="utf-8-sig")
            yaml_tekst, verwijderd = verwijder_ongeldige_tekens(ruwe_tekst)
            if verwijderd:
                print(
                    f"Waarschuwing: {bestand}: {verwijderd} ongeldig(e) "
                    "teken(s) verwijderd",
                    file=sys.stderr,
                )
            resultaten.append((bestand, yaml.safe_load(yaml_tekst)))
        except (OSError, UnicodeError, yaml.YAMLError) as fout:
            print(f"Waarschuwing: {bestand} overgeslagen: {fout}", file=sys.stderr)
    return resultaten


def normaliseer_kolomnaam(waarde: Any) -> str:
    return re.sub(r"[\s_-]+", "", str(waarde or "").strip().lower())


def lees_documentenlijst(excelpad: Path | None) -> dict[str, Scanverwijzing]:
    """Lees document_id, scanlocatie/spanlocatie en scannaam uit Excel."""
    if excelpad is None:
        return {}
    if not excelpad.is_file():
        raise FileNotFoundError(f"Documentenlijst bestaat niet: {excelpad}")

    werkmap = load_workbook(excelpad, read_only=True, data_only=True)
    try:
        werkblad = werkmap.active
        rijen = werkblad.iter_rows(values_only=True)
        try:
            koppen = next(rijen)
        except StopIteration as fout:
            raise ValueError("De documentenlijst is leeg") from fout

        kolommen = {
            normaliseer_kolomnaam(kop): index for index, kop in enumerate(koppen)
        }
        id_kolom = kolommen.get("documentid")
        locatie_kolom = kolommen.get("scanlocatie", kolommen.get("spanlocatie"))
        naam_kolom = kolommen.get("scannaam")
        if None in (id_kolom, locatie_kolom, naam_kolom):
            raise ValueError(
                "Excel moet document_id, scanlocatie/spanlocatie en scannaam bevatten"
            )

        verwijzingen: dict[str, Scanverwijzing] = {}
        for rij in rijen:
            document_id = normaliseer_id(rij[id_kolom])
            if document_id is None:
                continue
            locatie = rij[locatie_kolom]
            scannaam = rij[naam_kolom]
            if not heeft_waarde(locatie) or not heeft_waarde(scannaam):
                continue
            locatiepad = Path(str(locatie).strip()).expanduser()
            if not locatiepad.is_absolute():
                locatiepad = excelpad.parent / locatiepad
            verwijzingen[document_id] = Scanverwijzing(
                bronbestand=locatiepad / str(scannaam).strip(),
                originele_bestandsnaam=str(scannaam).strip(),
            )
        return verwijzingen
    finally:
        werkmap.close()


def agenda_item_kop(agenda_item: dict[str, Any]) -> tuple[Any, Any]:
    type_data = agenda_item.get("type_data")
    if not isinstance(type_data, dict):
        return None, None
    return (
        type_data.get("title_prefix"),
        waarde_met_id(type_data.get("attributes"), 1),
    )


def agenda_attribuut(agenda_item: dict[str, Any], attribuut_id: int) -> Any:
    type_data = agenda_item.get("type_data")
    attributen = type_data.get("attributes") if isinstance(type_data, dict) else None
    return waarde_met_id(attributen, attribuut_id)


def zichtbaar_agenda_item(item: dict[str, Any]) -> bool:
    prefix, naam = agenda_item_kop(item)
    return heeft_waarde(prefix) or heeft_waarde(naam)


def zichtbare_agenda_bomen(items: Any) -> list[dict[str, Any]]:
    """Behoud hiërarchie; promoveer kinderen van een leeg agenda-item."""
    resultaat: list[dict[str, Any]] = []
    if not isinstance(items, list):
        return resultaat
    for item in items:
        if not isinstance(item, dict):
            continue
        kinderen = zichtbare_agenda_bomen(item.get("agenda_items"))
        if zichtbaar_agenda_item(item):
            kopie = dict(item)
            kopie["_zichtbare_kinderen"] = kinderen
            resultaat.append(kopie)
        else:
            resultaat.extend(kinderen)
    return resultaat


def is_bestandsdocument(document: dict[str, Any]) -> bool:
    versions = document.get("versions")
    return isinstance(versions, list) and any(
        isinstance(versie, dict)
        and str(versie.get("type", "")).strip().lower() == "file"
        for versie in versions
    )


def eerste_bestandsnaam(document: dict[str, Any]) -> str | None:
    versions = document.get("versions")
    if not isinstance(versions, list):
        return None
    for versie in versions:
        if (
            isinstance(versie, dict)
            and str(versie.get("type", "")).strip().lower() == "file"
            and heeft_waarde(versie.get("file_name"))
        ):
            return Path(str(versie["file_name"])).name
    return None


def bestandsnaam_voor_document(
    document: dict[str, Any], verwijzing: Scanverwijzing | None
) -> str:
    if verwijzing is not None and heeft_waarde(verwijzing.originele_bestandsnaam):
        return Path(verwijzing.originele_bestandsnaam).name
    yaml_naam = eerste_bestandsnaam(document)
    if yaml_naam:
        return yaml_naam
    url = document.get("url")
    if heeft_waarde(url):
        url_naam = Path(unquote(urlparse(str(url)).path)).name
        if url_naam:
            return url_naam
    return "document"


def download_document(url: str, doelbestand: Path) -> None:
    parsed = urlparse(url)
    if parsed.scheme.lower() not in {"http", "https"}:
        raise ValueError(f"niet-ondersteund URL-schema: {parsed.scheme or 'ontbreekt'}")
    verzoek = Request(url, headers={"User-Agent": "Maak-ORI-A-YML-import/1.0"})
    tijdelijk_pad: Path | None = None
    try:
        with urlopen(verzoek, timeout=DOWNLOAD_TIMEOUT_SECONDEN) as antwoord:
            definitieve_url = antwoord.geturl()
            if urlparse(definitieve_url).scheme.lower() not in {"http", "https"}:
                raise ValueError("download verwees door naar een ongeldige URL")
            with tempfile.NamedTemporaryFile(
                mode="wb", dir=doelbestand.parent, prefix=".download-", delete=False
            ) as tijdelijk:
                tijdelijk_pad = Path(tijdelijk.name)
                shutil.copyfileobj(antwoord, tijdelijk)
        tijdelijk_pad.replace(doelbestand)
        tijdelijk_pad = None
    finally:
        if tijdelijk_pad is not None:
            tijdelijk_pad.unlink(missing_ok=True)


def qnaam(naam: str) -> str:
    return f"{{{ORIA_NS}}}{naam}"


def element(parent: ET.Element, naam: str, waarde: Any = None) -> ET.Element:
    kind = ET.SubElement(parent, qnaam(naam))
    if waarde is not None:
        kind.text = str(waarde)
    return kind


def voeg_verwijzing_toe(parent: ET.Element, verwijzing_id: str, naam: str | None) -> None:
    element(parent, "verwijzingID", verwijzing_id)
    if heeft_waarde(naam):
        element(parent, "verwijzingNaam", str(naam).strip())


def voeg_begrip_toe(
    parent: ET.Element,
    label: str,
    lijst_id: str,
    lijst_naam: str | None = None,
    code: str | None = None,
) -> None:
    element(parent, "begripLabel", label)
    if heeft_waarde(code):
        element(parent, "begripCode", code)
    lijst = element(parent, "verwijzingBegrippenlijst")
    voeg_verwijzing_toe(lijst, lijst_id, lijst_naam)


def voeg_informatieobject_toe(
    parent: ET.Element,
    elementnaam: str,
    object_id: str,
    objectnaam: str,
    type_label: str | None,
    type_lijst: str,
) -> ET.Element:
    informatieobject = element(parent, elementnaam)
    if heeft_waarde(type_label):
        soort = element(informatieobject, "informatieobjectType")
        voeg_begrip_toe(soort, str(type_label).strip(), type_lijst)
    verwijzing = element(informatieobject, "verwijzingInformatieobject")
    voeg_verwijzing_toe(verwijzing, object_id, objectnaam)
    return informatieobject


def xml_datetime(waarde: Any) -> str | None:
    if not heeft_waarde(waarde):
        return None
    tekst = str(waarde).strip().replace(" ", "T", 1)
    try:
        return datetime.fromisoformat(tekst).isoformat()
    except ValueError:
        return None


def xml_date(waarde: Any) -> str | None:
    tijdstip = xml_datetime(waarde)
    return tijdstip[:10] if tijdstip else None


def tijdstip_met_offset(start_date: Any, offset: Any) -> str | None:
    basis = xml_datetime(start_date)
    seconden = niet_negatief_getal(offset)
    if basis is None or seconden is None:
        return None
    return (datetime.fromisoformat(basis) + timedelta(seconds=seconden)).isoformat()


def niet_negatief_getal(waarde: Any) -> int | None:
    """Geef een ORI-A-geschikt aantal seconden of None terug."""
    if waarde in (None, ""):
        return None
    try:
        getal = int(waarde)
    except (TypeError, ValueError):
        return None
    return getal if getal >= 0 else None


def media_documenten(meeting: dict[str, Any]) -> list[dict[str, Any]]:
    """Zet Notubiz audio/video om naar documentachtige mappings."""
    meeting_id = normaliseer_id(meeting.get("id")) or "onbekend"
    media = meeting.get("media")
    resultaat: list[dict[str, Any]] = []
    if not isinstance(media, dict):
        return resultaat
    for mediatype in ("audio", "video"):
        items = media.get(mediatype)
        if not isinstance(items, list):
            continue
        for volgorde, item in enumerate(items, start=1):
            if not isinstance(item, dict) or not heeft_waarde(item.get("filename")):
                continue
            url = item.get("download")
            if heeft_waarde(url) and str(url).startswith("//"):
                url = "https:" + str(url)
            resultaat.append(
                {
                    "id": f"{meeting_id}-media-{mediatype}-{volgorde}",
                    "title": str(item["filename"]).strip(),
                    "url": url,
                    "types": [{"value": mediatype.capitalize()}],
                    "versions": [
                        {
                            "type": "file",
                            "file_name": str(item["filename"]).strip(),
                        }
                    ],
                    "_media_type": mediatype.capitalize(),
                }
            )
    return resultaat


def ori_document_id(document: dict[str, Any]) -> str:
    return f"document-{normaliseer_id(document.get('id')) or 'onbekend'}"


def document_naam(document: dict[str, Any]) -> str:
    titel = document.get("title")
    return str(titel).strip() if heeft_waarde(titel) else eerste_bestandsnaam(document) or "Document"


def bouw_agendapunt(
    parent: ET.Element,
    agenda_item: dict[str, Any],
    meeting_id: str,
    start_date: Any,
    media: list[dict[str, Any]],
    documenten_voor_overdracht: list[dict[str, Any]],
) -> None:
    agenda_id = normaliseer_id(agenda_item.get("id")) or "onbekend"
    prefix, naam = agenda_item_kop(agenda_item)
    naamtekst = str(naam).strip() if heeft_waarde(naam) else str(prefix).strip()
    agendapunt = element(parent, "agendapunt" if parent.tag == qnaam("ORI-A") else "heeftAlsSubagendapunt")
    element(agendapunt, "ID", f"agendapunt-{agenda_id}")
    element(agendapunt, "naam", naamtekst)

    if heeft_waarde(prefix):
        weergave = str(prefix).strip()
        nummer = weergave.rstrip(".").strip()
        if re.match(r"^\d+", nummer):
            element(agendapunt, "geplandVolgnummer", nummer)
            element(agendapunt, "volgnummer", nummer)
        element(agendapunt, "volgnummerWeergave", weergave)

    omschrijving = html_naar_tekst(agenda_attribuut(agenda_item, 3))
    if omschrijving:
        element(agendapunt, "omschrijving", omschrijving)

    starttijd = tijdstip_met_offset(start_date, agenda_item.get("start_offset"))
    eindtijd = tijdstip_met_offset(start_date, agenda_item.get("end_offset"))
    if starttijd:
        element(agendapunt, "starttijd", starttijd)
    if eindtijd:
        element(agendapunt, "eindtijd", eindtijd)

    start_offset = niet_negatief_getal(agenda_item.get("start_offset"))
    eind_offset = niet_negatief_getal(agenda_item.get("end_offset"))
    if start_offset is not None and media:
        for mediabron in media:
            tijdsaanduiding = element(agendapunt, "tijdsaanduidingMediabron")
            element(tijdsaanduiding, "aanvang", start_offset)
            if eind_offset is not None:
                element(tijdsaanduiding, "einde", eind_offset)
            if mediabron is not None and len(media) > 1:
                voeg_informatieobject_toe(
                    tijdsaanduiding,
                    "isRelatiefTot",
                    ori_document_id(mediabron),
                    document_naam(mediabron),
                    mediabron.get("_media_type"),
                    MEDIABRONTYPEN,
                )

    vergadering = element(agendapunt, "wordtBehandeldTijdens")
    voeg_verwijzing_toe(vergadering, f"vergadering-{meeting_id}", None)

    documenten = agenda_item.get("documents")
    if isinstance(documenten, list):
        for document in documenten:
            if not isinstance(document, dict) or not is_bestandsdocument(document):
                continue
            documenten_voor_overdracht.append(document)
            categorie = eerste_gevulde_waarde(document.get("types"))
            voeg_informatieobject_toe(
                agendapunt,
                "heeftAlsBijlage",
                ori_document_id(document),
                document_naam(document),
                str(categorie).strip() if heeft_waarde(categorie) else "Document",
                VERGADERSTUKTYPEN,
            )

    for kind in agenda_item.get("_zichtbare_kinderen", []):
        bouw_agendapunt(
            agendapunt,
            kind,
            meeting_id,
            start_date,
            media,
            documenten_voor_overdracht,
        )


def bouw_ori_xml(meeting: dict[str, Any]) -> tuple[ET.ElementTree, list[dict[str, Any]]]:
    """Bouw één volledig ORI-A-document voor een Notubiz-meeting."""
    meeting_id = normaliseer_id(meeting.get("id"))
    meetingnaam = waarde_met_id(meeting.get("attributes"), 1)
    meetingnaam = str(meetingnaam).strip() if heeft_waarde(meetingnaam) else ""
    planningen = meeting.get("plannings")
    planning = planningen[0] if isinstance(planningen, list) and planningen else {}
    start_date = planning.get("start_date") if isinstance(planning, dict) else None
    datum = xml_date(start_date)
    agenda_bomen = zichtbare_agenda_bomen(meeting.get("agenda_items"))
    if meeting_id is None or not meetingnaam or datum is None:
        raise ValueError("meeting mist id, naam of een geldige start_date")
    if not agenda_bomen:
        raise ValueError("meeting bevat geen benoemde agenda-items")

    root = ET.Element(qnaam("ORI-A"))
    root.set(qnaam("schemaLocation").replace(ORIA_NS, XSI_NS), SCHEMA_LOCATIE)
    vergadering = element(root, "vergadering")
    element(vergadering, "ID", f"vergadering-{meeting_id}")
    element(vergadering, "naam", meetingnaam)
    element(vergadering, "geplandeDatum", datum)
    element(vergadering, "datum", datum)
    geplande_aanvang = xml_datetime(start_date)
    if geplande_aanvang:
        element(vergadering, "geplandeAanvang", geplande_aanvang)
    gepland_einde = xml_datetime(planning.get("end_date")) if isinstance(planning, dict) else None
    if gepland_einde:
        element(vergadering, "geplandEinde", gepland_einde)
    publicatiedatum = xml_datetime(meeting.get("last_modified"))
    if publicatiedatum:
        element(vergadering, "publicatiedatum", publicatiedatum)

    vergaderingstype = "Raadsvergadering" if "raad" in meetingnaam.lower() else meetingnaam
    soort = element(vergadering, "type")
    voeg_begrip_toe(soort, vergaderingstype, VERGADERINGSTYPEN)
    toelichting = html_naar_tekst(waarde_met_id(meeting.get("attributes"), 3))
    if toelichting:
        element(vergadering, "toelichting", toelichting)

    gremium = meeting.get("gremium")
    gremium_element = element(vergadering, "georganiseerdDoorGremium")
    element(gremium_element, "naam", meetingnaam)
    if isinstance(gremium, dict) and normaliseer_id(gremium.get("id")):
        element(gremium_element, "identificatie", normaliseer_id(gremium.get("id")))

    locatie = waarde_met_id(meeting.get("attributes"), 50)
    if heeft_waarde(locatie):
        element(vergadering, "locatie", str(locatie).strip())
    if heeft_waarde(meeting.get("url")):
        element(vergadering, "weblocatie", str(meeting["url"]).strip())
    element(vergadering, "status", "Geannuleerd" if meeting.get("canceled") else "Gehouden")

    overheidsorgaan = element(vergadering, "overheidsorgaan")
    voeg_begrip_toe(
        overheidsorgaan,
        "Gemeente Stichtse Vecht",
        GEMEENTENREGISTER,
        "Register gemeenten compleet",
        "gm1904",
    )

    media = media_documenten(meeting)
    for mediabron in media:
        voeg_informatieobject_toe(
            vergadering,
            "isVastgelegdMiddels",
            ori_document_id(mediabron),
            document_naam(mediabron),
            mediabron.get("_media_type"),
            MEDIABRONTYPEN,
        )

    documenten_voor_overdracht: list[dict[str, Any]] = list(media)
    for agenda_item in agenda_bomen:
        bouw_agendapunt(
            root,
            agenda_item,
            meeting_id,
            start_date,
            media,
            documenten_voor_overdracht,
        )
    return ET.ElementTree(root), documenten_voor_overdracht


def schrijf_ori_xml(boom: ET.ElementTree, doelbestand: Path) -> None:
    ET.indent(boom, space="    ")
    boom.write(doelbestand, encoding="utf-8", xml_declaration=True)


def valideer_xml(xmlbestand: Path, xsd_pad: Path) -> None:
    """Valideer met lxml als een XSD-pad is opgegeven."""
    try:
        from lxml import etree
    except ImportError as fout:
        raise RuntimeError("voor --xsd is het pakket lxml nodig") from fout
    schema = etree.XMLSchema(etree.parse(str(xsd_pad)))
    document = etree.parse(str(xmlbestand))
    schema.assertValid(document)


def verwerk_documentbestand(
    document: dict[str, Any],
    documentenlijst: dict[str, Scanverwijzing],
    meetingmap: Path,
    nummering: Meetingnummering,
    statistieken: Statistieken,
) -> None:
    document_id = normaliseer_id(document.get("id"))
    if document_id is None or document_id in statistieken.verwerkte_document_ids:
        return
    statistieken.verwerkte_document_ids.add(document_id)
    verwijzing = documentenlijst.get(document_id)
    originele_naam = bestandsnaam_voor_document(document, verwijzing)
    # Reik het nummer pas definitief uit na een geslaagde kopie of download.
    objectnaam = nummering.kandidaat()
    documentmap = meetingmap / objectnaam
    extensie = Path(originele_naam).suffix
    doelbestand = documentmap / f"{objectnaam}{extensie}"
    documentmap.mkdir(parents=True, exist_ok=False)

    if verwijzing is not None and verwijzing.bronbestand.is_file():
        try:
            shutil.copy2(verwijzing.bronbestand, doelbestand)
        except OSError as fout:
            print(
                f"Waarschuwing: document {document_id} kon niet worden gekopieerd: "
                f"{fout}; download wordt geprobeerd",
                file=sys.stderr,
            )
        else:
            nummering.nieuw()
            statistieken.gekopieerd += 1
            return

    url = document.get("url")
    if not heeft_waarde(url):
        print(
            f"Waarschuwing: document {document_id} heeft geen bruikbare scan of URL",
            file=sys.stderr,
        )
        shutil.rmtree(documentmap)
        statistieken.overgeslagen += 1
        return
    try:
        download_document(str(url).strip(), doelbestand)
    except (OSError, ValueError) as fout:
        print(
            f"Waarschuwing: document {document_id} kon niet worden gedownload: {fout}",
            file=sys.stderr,
        )
        shutil.rmtree(documentmap)
        statistieken.overgeslagen += 1
        return
    nummering.nieuw()
    statistieken.gedownload += 1


def uitvoermap_is_niet_leeg(uitvoermap: Path) -> bool:
    return uitvoermap.is_dir() and next(uitvoermap.iterdir(), None) is not None


def verwerk_ymls(
    bronmap: Path,
    uitvoermap: Path,
    documentenlijstpad: Path | None,
    xsd_pad: Path | None,
    alleen_xml: bool,
    overschrijf: bool,
    prefix: str,
) -> int:
    if uitvoermap_is_niet_leeg(uitvoermap) and not overschrijf:
        print(
            f"Waarschuwing: uitvoermap {uitvoermap} is niet leeg. Gebruik "
            "--overschrijf om toch nieuwe bestanden toe te voegen.",
            file=sys.stderr,
        )
        return 5
    if xsd_pad is not None and not xsd_pad.is_file():
        print(f"Fout: XSD bestaat niet: {xsd_pad}", file=sys.stderr)
        return 6
    try:
        documentenlijst = lees_documentenlijst(documentenlijstpad)
        meetings = lees_yml_bestanden(bronmap)
    except (OSError, ValueError) as fout:
        print(f"Fout: {fout}", file=sys.stderr)
        return 2
    if not meetings:
        print(f"Geen leesbare .yml-bestanden gevonden in {bronmap}", file=sys.stderr)
        return 1

    statistieken = Statistieken()
    uitvoermap.mkdir(parents=True, exist_ok=True)
    meetingvolgnummer = 0
    for bronbestand, meeting in meetings:
        kandidaat_meetingnummer = meetingvolgnummer + 1
        relatief = bronbestand.relative_to(bronmap)
        meetingmap = (
            uitvoermap
            / relatief.parent
            / f"{prefix}_{kandidaat_meetingnummer:04d}"
        )
        nummering = Meetingnummering(kandidaat_meetingnummer, prefix)
        tijdelijk_xmlbestand: Path | None = None
        try:
            boom, documenten = bouw_ori_xml(meeting)
            meetingmap.mkdir(parents=True, exist_ok=True)
            xmlnaam = nummering.nieuw()
            xmlbestand = meetingmap / f"{xmlnaam}.ori-a.xml"
            tijdelijk_xmlbestand = meetingmap / f".{xmlnaam}.ori-a.xml.tmp"
            schrijf_ori_xml(boom, tijdelijk_xmlbestand)
            if xsd_pad is not None:
                valideer_xml(tijdelijk_xmlbestand, xsd_pad)
            tijdelijk_xmlbestand.replace(xmlbestand)
            tijdelijk_xmlbestand = None
        except (OSError, TypeError, ValueError, RuntimeError) as fout:
            if tijdelijk_xmlbestand is not None:
                tijdelijk_xmlbestand.unlink(missing_ok=True)
            print(f"Waarschuwing: {relatief} overgeslagen: {fout}", file=sys.stderr)
            continue
        except Exception as fout:
            # lxml.etree.DocumentInvalid is niet beschikbaar zonder lxml-import.
            if tijdelijk_xmlbestand is not None:
                tijdelijk_xmlbestand.unlink(missing_ok=True)
            print(f"Waarschuwing: ORI-A-validatie mislukte voor {relatief}: {fout}", file=sys.stderr)
            continue

        meetingvolgnummer = kandidaat_meetingnummer
        statistieken.xml_bestanden += 1
        statistieken.verwerkte_document_ids.clear()
        print(f"ORI-A geschreven: {xmlbestand}")
        if alleen_xml:
            continue
        for document in documenten:
            verwerk_documentbestand(
                document, documentenlijst, meetingmap, nummering, statistieken
            )

    print(
        f"{statistieken.xml_bestanden} ORI-A-bestand(en), "
        f"{statistieken.gekopieerd} bestand(en) gekopieerd, "
        f"{statistieken.gedownload} gedownload en "
        f"{statistieken.overgeslagen} overgeslagen."
    )
    return 0 if statistieken.xml_bestanden else 3


def parse_parameters() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Lees Notubiz-YAML, genereer ORI-A XML en verzamel documenten/media."
        )
    )
    parser.add_argument(
        "--bronmap",
        type=Path,
        default=Path(os.environ.get("YML_BRONMAP", "input")),
        help="Hoofdmap met .yml-bestanden; submappen worden meegenomen",
    )
    parser.add_argument(
        "--uitvoermap",
        type=Path,
        default=Path(os.environ.get("YML_UITVOERMAP", "output")),
        help="Hoofdmap voor ORI-A XML en verzamelde bestanden",
    )
    documentenlijst = os.environ.get("YML_DOCUMENTENLIJST")
    parser.add_argument(
        "--documentenlijst",
        type=Path,
        default=Path(documentenlijst) if documentenlijst else None,
        help="Optionele Excel met document_id, scanlocatie/spanlocatie en scannaam",
    )
    xsd = os.environ.get("ORIA_XSD")
    parser.add_argument(
        "--xsd",
        type=Path,
        default=Path(xsd) if xsd else None,
        help="Optioneel ORI-A.xsd-pad voor validatie van iedere gegenereerde XML",
    )
    parser.add_argument(
        "--alleen-xml",
        action="store_true",
        help="Genereer en valideer XML zonder documenten of media over te dragen",
    )
    parser.add_argument(
        "--overschrijf",
        action="store_true",
        help="Sta verwerking toe wanneer de uitvoermap niet leeg is",
    )
    parser.add_argument(
        "--prefix",
        default=os.environ.get("ORIA_PREFIX", STANDAARD_PREFIX),
        help=f"Prefix voor uitvoernamen (standaard: {STANDAARD_PREFIX})",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_parameters()
    return verwerk_ymls(
        args.bronmap,
        args.uitvoermap,
        args.documentenlijst,
        args.xsd,
        args.alleen_xml,
        args.overschrijf,
        args.prefix,
    )


if __name__ == "__main__":
    raise SystemExit(main())
